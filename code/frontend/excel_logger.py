"""
excel_logger.py – Background Excel Logger

Periodically writes the sensor history to a temporary Excel workbook
at  code/database/sensor_readings_temp.xlsx

Columns:
  A: Timestamp  (HH:MM:SS.mmm)
  B: Sensor Reading (%)
"""

import threading
import time
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    print("[ExcelLogger] WARNING: openpyxl not installed – Excel logging disabled.")


_HEADER_FILL = "1a1a2e"
_ACCENT_FILL = "e94560"
_ROW_FILL_A  = "16213e"
_ROW_FILL_B  = "1e2d45"


class ExcelLogger:
    """Writes sensor data to Excel every `interval` seconds in a daemon thread."""

    INTERVAL_S = 2.0

    def __init__(self, sensor_state, output_path: str):
        self.sensor   = sensor_state
        self.path     = output_path
        self._stop    = threading.Event()
        self._thread  = threading.Thread(target=self._run, daemon=True)

    def start(self) -> None:
        """Start the background logging thread."""
        if _HAS_OPENPYXL:
            self._thread.start()
        else:
            print("[ExcelLogger] Skipping start – openpyxl unavailable.")

    def stop(self) -> None:
        """Signal the logging thread to stop and do a final write."""
        self._stop.set()

    # ------------------------------------------------------------------
    def _run(self) -> None:
        while not self._stop.is_set():
            time.sleep(self.INTERVAL_S)
            self._write()

    # ------------------------------------------------------------------
    def _write(self) -> None:
        history = self.sensor.get_history()
        if not history:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sensor Readings"

        # ── Header row ──────────────────────────────────────────────────
        headers = ["Timestamp", "Sensor Reading (%)"]
        header_font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        header_fill = PatternFill("solid", fgColor=_ACCENT_FILL)
        center      = Alignment(horizontal="center", vertical="center")
        thin_side   = Side(style="thin", color="2a2a4a")
        border      = Border(left=thin_side, right=thin_side,
                             top=thin_side, bottom=thin_side)

        for col, text in enumerate(headers, start=1):
            cell           = ws.cell(row=1, column=col, value=text)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border

        ws.row_dimensions[1].height = 22

        # ── Data rows ───────────────────────────────────────────────────
        fill_a = PatternFill("solid", fgColor=_ROW_FILL_A)
        fill_b = PatternFill("solid", fgColor=_ROW_FILL_B)
        data_font = Font(name="Segoe UI", size=9, color="EAEAEA")

        for row_idx, (ts, val) in enumerate(history, start=2):
            ts_str  = ts.strftime("%H:%M:%S.") + f"{ts.microsecond // 1000:03d}"
            row_fill = fill_a if row_idx % 2 == 0 else fill_b

            for col, value in enumerate([ts_str, round(val, 2)], start=1):
                cell           = ws.cell(row=row_idx, column=col, value=value)
                cell.font      = data_font
                cell.fill      = row_fill
                cell.alignment = center
                cell.border    = border

        # ── Column widths ────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 22

        # ── Meta sheet ──────────────────────────────────────────────────
        meta = wb.create_sheet("Info")
        meta["A1"] = "FloodGuard Sensor Log"
        meta["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        meta["A3"] = f"Total readings: {len(history)}"
        meta["A1"].font = Font(bold=True, size=12, name="Segoe UI")

        # ── Save ─────────────────────────────────────────────────────────
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        try:
            wb.save(self.path)
        except PermissionError:
            pass   # file may be open in Excel; skip this cycle
